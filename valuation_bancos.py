# -*- coding: utf-8 -*-
import sys, io as _io
sys.stdout = _io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ══════════════════════════════════════════════════════════════════════════
#  Valuation — Setor Bancário Brasileiro
#  Modelo: Excess Return (ROE vs Ke) — Damodaran
#  Bancos: Itaú, Bradesco, Banco do Brasil, Santander BR
#  Fontes: CVM (DFP), BCB (SELIC), yfinance (preços)
# ══════════════════════════════════════════════════════════════════════════

import os

import zipfile

import requests
import numpy as np
import pandas as pd
import yfinance as yf
import statsmodels.api as sm

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import io  # noqa: F811 — re-import after sys.stdout wrap


# ──────────────────────────────────────────────────────────────────
# Configuração dos bancos analisados e parâmetros globais do projeto
# ──────────────────────────────────────────────────────────────────

BANCOS = [
    {
        "nome":   "Itaú Unibanco",
        "ticker": "ITUB4.SA",
        "cnpj":   "60.872.504/0001-23",
    },
    {
        "nome":   "Bradesco",
        "ticker": "BBDC4.SA",
        "cnpj":   "60.746.948/0001-12",
    },
    {
        "nome":   "Banco do Brasil",
        "ticker": "BBAS3.SA",
        "cnpj":   "00.000.000/0001-91",
    },
    {
        "nome":   "Santander BR",
        "ticker": "SANB11.SA",
        "cnpj":   "90.400.888/0001-42",
    },
]

# Período de análise
# A CVM disponibiliza DFPs a partir de 2005 no formato aberto.
# O código tenta baixar desde ANO_INICIAL e pula anos indisponíveis.
ANO_INICIAL = 2010
ANO_FINAL   = 2024

# Pasta de saída (planilha Excel)
PASTA_SAIDA = r"C:\Users\adria\OneDrive\Área de Trabalho\ARQUIVOS HENRIQUE\antigravity\Valuation\bancos"

# ──────────────────────────────────────────────────────────────────
# 1. DOWNLOAD CVM
# ──────────────────────────────────────────────────────────────────

def baixar_dfp_zip(ano: int):
    """Baixa o ZIP da CVM para o ano. Retorna ZipFile ou None se indisponível."""
    url = (
        f"https://dados.cvm.gov.br/dados/CIA_ABERTA/DOC/DFP/DADOS/"
        f"dfp_cia_aberta_{ano}.zip"
    )
    try:
        resp = requests.get(url, timeout=90)
        resp.raise_for_status()
        return zipfile.ZipFile(io.BytesIO(resp.content))
    except Exception as e:
        print(f"  [AVISO] DFP {ano} indisponivel: {e}")
        return None


def extrair_csv_zip(zip_obj, prefixo: str) -> pd.DataFrame:
    """Extrai o CSV consolidado (con) para o prefixo fornecido (DRE ou BPP)."""
    candidatos = [
        n for n in zip_obj.namelist()
        if prefixo in n and "con" in n.lower()
    ]
    if not candidatos:
        return pd.DataFrame()
    return pd.read_csv(zip_obj.open(candidatos[0]), sep=";", encoding="latin1")


# ──────────────────────────────────────────────────────────────────
# 2. LOOP ÚNICO: baixa DRE + BPP de cada ZIP (1 request/ano)
# ──────────────────────────────────────────────────────────────────

def baixar_dados_cvm(anos) -> tuple[pd.DataFrame, pd.DataFrame, list]:
    lista_dre, lista_bpp, anos_ok = [], [], []

    for ano in anos:
        print(f"  [CVM] Baixando DFP {ano}...")
        zip_obj = baixar_dfp_zip(ano)
        if zip_obj is None:
            continue

        dre = extrair_csv_zip(zip_obj, "DRE")
        bpp = extrair_csv_zip(zip_obj, "BPP")

        if not dre.empty:
            lista_dre.append(dre)
        if not bpp.empty:
            lista_bpp.append(bpp)

        anos_ok.append(ano)

    dre_total = pd.concat(lista_dre, ignore_index=True) if lista_dre else pd.DataFrame()
    bpp_total = pd.concat(lista_bpp, ignore_index=True) if lista_bpp else pd.DataFrame()
    return dre_total, bpp_total, anos_ok


# ──────────────────────────────────────────────────────────────────
# 3. LUCRO LÍQUIDO DO CONTROLADOR (DRE)
# ──────────────────────────────────────────────────────────────────

def extrair_lucro(dre: pd.DataFrame, cnpj: str) -> pd.DataFrame:
    banco = dre[dre["CNPJ_CIA"] == cnpj]

    df = banco[
        banco["DS_CONTA"].str.contains("Lucro/Prejuízo", case=False, na=False)
        & (banco["ORDEM_EXERC"] == "ÚLTIMO")
    ].copy()

    df["Ano"] = pd.to_datetime(df["DT_REFER"]).dt.year
    df["Lucro_Liquido"] = df["VL_CONTA"] * 1_000

    return (
        df[["Ano", "Lucro_Liquido"]]
        .groupby("Ano", as_index=False)
        .sum()
        .sort_values("Ano")
    )

# ──────────────────────────────────────────────────────────────────
# 4. PATRIMÔNIO LÍQUIDO DO CONTROLADOR (BPP)
# ──────────────────────────────────────────────────────────────────

def extrair_pl(bpp: pd.DataFrame, cnpj: str) -> pd.DataFrame:
    banco = bpp[bpp["CNPJ_CIA"] == cnpj]
    df = banco[
        (banco["CD_CONTA"].isin(["2.08", "2.08.09"]))
        & (banco["ORDEM_EXERC"] == "ÚLTIMO")
    ].copy()
    df["Ano"] = pd.to_datetime(df["DT_REFER"]).dt.year
    df["VL_CONTA"] = df["VL_CONTA"] * 1_000       # MIL → R$

    pivot = df.pivot_table(
        index="Ano", columns="CD_CONTA", values="VL_CONTA", aggfunc="sum"
    )
    # PL Total (2.08) menos Participações de Não-Controladores (2.08.09)
    pivot["PL_Controlador"] = (
        pivot.get("2.08", pd.Series(0, index=pivot.index))
        - pivot.get("2.08.09", pd.Series(0, index=pivot.index))
    )
    return (
        pivot[["PL_Controlador"]]
        .reset_index()
        .sort_values("Ano")
        .reset_index(drop=True)
    )


# ──────────────────────────────────────────────────────────────────
# 5. ROE  (Lucro_t / PL_(t-1))
# ──────────────────────────────────────────────────────────────────

def calcular_roe(df_lucro: pd.DataFrame, df_pl: pd.DataFrame) -> pd.DataFrame:
    df = pd.merge(df_lucro, df_pl, on="Ano", how="inner")
    df["PL_Inicial"] = df["PL_Controlador"].shift(1)
    df["ROE"] = df["Lucro_Liquido"] / df["PL_Inicial"]
    return df.dropna().reset_index(drop=True)


# ──────────────────────────────────────────────────────────────────
# 6. BETA via OLS (retornos mensais)
# ──────────────────────────────────────────────────────────────────

def calcular_beta(ticker: str, ano_ini: int, ano_fim: int) -> float:
    precos = yf.download(
        [ticker, "^BVSP"],
        start=f"{ano_ini}-01-01",
        end=f"{ano_fim}-12-31",
        auto_adjust=True,
        progress=False,
    )["Close"]
    mensais = precos.resample("ME").last()
    ret = mensais.pct_change().dropna()
    ret.columns = ["ACAO", "IBOV"]
    X = sm.add_constant(ret["IBOV"])
    modelo = sm.OLS(ret["ACAO"], X).fit()
    return float(modelo.params["IBOV"])


# ──────────────────────────────────────────────────────────────────
# 7. SELIC ANUAL (BCB, série 11 — taxa Over/Selic diária)
# ──────────────────────────────────────────────────────────────────

def baixar_selic(ano_ini: int, ano_fim: int) -> pd.DataFrame:
    SERIE = 11
    todos = []
    # A API do BCB tem limite de ~10 anos por consulta
    for inicio in range(ano_ini, ano_fim + 1, 10):
        fim_bloco = min(inicio + 9, ano_fim)
        url = (
            f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{SERIE}/dados"
            f"?formato=json&dataInicial=01/01/{inicio}&dataFinal=31/12/{fim_bloco}"
        )
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        todos.extend(resp.json())

    df = pd.DataFrame(todos)
    df["data"] = pd.to_datetime(df["data"], dayfirst=True)
    df["valor"] = pd.to_numeric(df["valor"])
    df["ano"] = df["data"].dt.year
    df["fator_diario"] = 1 + df["valor"] / 100

    anual = df.groupby("ano")["fator_diario"].prod().reset_index()
    anual["selic_anual"] = anual["fator_diario"] - 1
    return anual[["ano", "selic_anual"]]


# ──────────────────────────────────────────────────────────────────
# 8. RETORNO ANUAL DO IBOV
# ──────────────────────────────────────────────────────────────────

def baixar_ibov_anual(ano_ini: int, ano_fim: int) -> pd.DataFrame:
    ibov = yf.download(
        "^BVSP",
        start=f"{ano_ini}-01-01",
        end=f"{ano_fim}-12-31",
        auto_adjust=False,
        progress=False,
    )
    if isinstance(ibov.columns, pd.MultiIndex):
        ibov.columns = ibov.columns.get_level_values(0)

    ibov = ibov[["Adj Close"]].reset_index()
    ibov["ano"] = ibov["Date"].dt.year

    anual = (
        ibov.groupby("ano")
        .agg(preco_inicial=("Adj Close", "first"), preco_final=("Adj Close", "last"))
        .reset_index()
    )
    anual["retorno_ibov"] = anual["preco_final"] / anual["preco_inicial"] - 1
    return anual[["ano", "retorno_ibov"]]


# ──────────────────────────────────────────────────────────────────
# 9. CAPM + SPREAD DE VALOR
# ──────────────────────────────────────────────────────────────────

def calcular_capm(
    df_roe: pd.DataFrame,
    df_selic: pd.DataFrame,
    df_ibov: pd.DataFrame,
    beta: float,
) -> pd.DataFrame:
    df = (
        df_roe
        .merge(df_selic, left_on="Ano", right_on="ano")
        .merge(df_ibov, left_on="Ano", right_on="ano")
        .rename(columns={"selic_anual": "Rf", "retorno_ibov": "Rm"})
    )
    df["beta"] = beta
    df["premio_risco"] = df["Rm"] - df["Rf"]
    df["Ke"] = df["Rf"] + beta * df["premio_risco"]
    df["spread_valor"] = df["ROE"] - df["Ke"]

    colunas = [
        "Ano", "Lucro_Liquido", "PL_Controlador", "PL_Inicial",
        "ROE", "beta", "Rf", "Rm", "Ke", "spread_valor",
    ]
    return df[colunas].reset_index(drop=True)


# ──────────────────────────────────────────────────────────────────
# 10. GERAÇÃO DO EXCEL COM GRÁFICOS
# ──────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="FF1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=10)
TITLE_FONT  = Font(bold=True, size=13, color="FF1E3A5F")
COL_PCTS    = {"ROE", "Rf", "Rm", "Ke", "spread_valor"}
COL_BRL     = {"Lucro_Liquido", "PL_Controlador", "PL_Inicial"}


def _escrever_df(ws, df: pd.DataFrame, titulo: str | None = None) -> tuple[int, int]:
    """Escreve DataFrame na worksheet. Retorna (linha_header, última_linha)."""
    row_offset = 1
    if titulo:
        cell = ws.cell(1, 1, titulo)
        cell.font = TITLE_FONT
        row_offset = 3

    cols = list(df.columns)
    # Cabeçalho
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row_offset, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for r, row in enumerate(df.itertuples(index=False), row_offset + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            col_name = cols[c - 1]
            if col_name in COL_PCTS and isinstance(val, float):
                cell.number_format = "0.00%"
            elif col_name in COL_BRL and isinstance(val, (int, float)):
                cell.number_format = '#,##0'

    last_row = row_offset + len(df)

    # Largura das colunas
    for c, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = max(len(col) + 4, 14)

    return row_offset, last_row


def _tabela_pivot(df_todos, campo, bancos_unicos, anos_unicos, ws):
    """Preenche worksheet em formato pivot (anos = linhas, bancos = colunas)."""
    ws.cell(1, 1, "Ano").font = HEADER_FONT
    ws.cell(1, 1).fill = HEADER_FILL
    for c, b in enumerate(bancos_unicos, 2):
        cell = ws.cell(1, c, b)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for r, ano in enumerate(anos_unicos, 2):
        ws.cell(r, 1, ano)
        for c, banco in enumerate(bancos_unicos, 2):
            sub = df_todos[(df_todos["Banco"] == banco) & (df_todos["Ano"] == ano)][campo]
            val = float(sub.values[0]) if len(sub) > 0 else None
            cell = ws.cell(r, c, val)
            if val is not None and campo in COL_PCTS:
                cell.number_format = "0.00%"

    for c in range(1, len(bancos_unicos) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 16


def criar_excel(resultados: dict, pasta_saida: str) -> str:
    os.makedirs(pasta_saida, exist_ok=True)
    caminho = os.path.join(pasta_saida, "valuation_bancos.xlsx")

    # DataFrame consolidado
    frames = []
    for nome, df in resultados.items():
        d = df.copy()
        d.insert(0, "Banco", nome)
        frames.append(d)
    df_todos = pd.concat(frames, ignore_index=True)

    bancos_unicos = [b["nome"] for b in BANCOS if b["nome"] in resultados]
    anos_unicos   = sorted(df_todos["Ano"].unique())
    n_anos        = len(anos_unicos)

    wb = Workbook()

    # ── ABA 1: Consolidado ───────────────────────────────────────
    ws_cons = wb.active
    ws_cons.title = "Setor Bancário"
    _escrever_df(ws_cons, df_todos, "Valuation — Setor Bancário Brasileiro (2005–2024)")

    # ── ABA 2: ROE Comparativo ────────────────────────────────────
    ws_roe = wb.create_sheet("ROE Comparativo")
    ws_roe.cell(1, 1).value  # trigger
    _tabela_pivot(df_todos, "ROE", bancos_unicos, anos_unicos, ws_roe)

    chart_roe = LineChart()
    chart_roe.title  = "ROE por Banco"
    chart_roe.style  = 10
    chart_roe.height = 14
    chart_roe.width  = 26
    chart_roe.y_axis.title   = "ROE"
    chart_roe.x_axis.title   = "Ano"
    chart_roe.y_axis.numFmt  = "0%"

    cats = Reference(ws_roe, min_col=1, min_row=2, max_row=n_anos + 1)
    for c in range(2, len(bancos_unicos) + 2):
        data = Reference(ws_roe, min_col=c, min_row=1, max_row=n_anos + 1)
        chart_roe.add_data(data, titles_from_data=True)
    chart_roe.set_categories(cats)
    ws_roe.add_chart(chart_roe, f"A{n_anos + 4}")

    # ── ABA 3: Spread de Valor Comparativo ───────────────────────
    ws_sp = wb.create_sheet("Spread de Valor")
    _tabela_pivot(df_todos, "spread_valor", bancos_unicos, anos_unicos, ws_sp)

    chart_sp = BarChart()
    chart_sp.type   = "col"
    chart_sp.title  = "Spread de Valor (ROE − Ke) por Banco"
    chart_sp.style  = 10
    chart_sp.height = 14
    chart_sp.width  = 26
    chart_sp.y_axis.title  = "Spread"
    chart_sp.y_axis.numFmt = "0%"

    cats2 = Reference(ws_sp, min_col=1, min_row=2, max_row=n_anos + 1)
    for c in range(2, len(bancos_unicos) + 2):
        data = Reference(ws_sp, min_col=c, min_row=1, max_row=n_anos + 1)
        chart_sp.add_data(data, titles_from_data=True)
    chart_sp.set_categories(cats2)
    ws_sp.add_chart(chart_sp, f"A{n_anos + 4}")

    # ── ABA 4: Ke Comparativo ─────────────────────────────────────
    ws_ke = wb.create_sheet("Ke Comparativo")
    _tabela_pivot(df_todos, "Ke", bancos_unicos, anos_unicos, ws_ke)

    chart_ke = LineChart()
    chart_ke.title  = "Custo de Capital (Ke) por Banco"
    chart_ke.style  = 10
    chart_ke.height = 14
    chart_ke.width  = 26
    chart_ke.y_axis.title  = "Ke"
    chart_ke.y_axis.numFmt = "0%"
    cats3 = Reference(ws_ke, min_col=1, min_row=2, max_row=n_anos + 1)
    for c in range(2, len(bancos_unicos) + 2):
        data = Reference(ws_ke, min_col=c, min_row=1, max_row=n_anos + 1)
        chart_ke.add_data(data, titles_from_data=True)
    chart_ke.set_categories(cats3)
    ws_ke.add_chart(chart_ke, f"A{n_anos + 4}")

    # ── ABAS POR BANCO ────────────────────────────────────────────
    for nome, df in resultados.items():
        ws = wb.create_sheet(nome[:31])
        row_ini, row_fim = _escrever_df(ws, df, f"Valuation — {nome}")

        cols = list(df.columns)
        col_ano    = cols.index("Ano")    + 1
        col_roe    = cols.index("ROE")    + 1
        col_ke     = cols.index("Ke")     + 1
        col_spread = cols.index("spread_valor") + 1

        cats_b = Reference(ws, min_col=col_ano, min_row=row_ini + 1, max_row=row_fim)

        # Gráfico 1: ROE vs Ke
        c1 = LineChart()
        c1.title  = f"{nome} — ROE vs Ke"
        c1.style  = 10
        c1.height = 14
        c1.width  = 26
        c1.y_axis.title  = "Taxa"
        c1.y_axis.numFmt = "0%"
        c1.x_axis.title  = "Ano"
        d_roe = Reference(ws, min_col=col_roe, min_row=row_ini, max_row=row_fim)
        d_ke  = Reference(ws, min_col=col_ke,  min_row=row_ini, max_row=row_fim)
        c1.add_data(d_roe, titles_from_data=True)
        c1.add_data(d_ke,  titles_from_data=True)
        c1.set_categories(cats_b)
        ws.add_chart(c1, f"A{row_fim + 3}")

        # Gráfico 2: Spread de Valor (barras)
        c2 = BarChart()
        c2.type   = "col"
        c2.title  = f"{nome} — Spread de Valor (ROE − Ke)"
        c2.style  = 10
        c2.height = 14
        c2.width  = 26
        c2.y_axis.title  = "Spread"
        c2.y_axis.numFmt = "0%"
        d_sp = Reference(ws, min_col=col_spread, min_row=row_ini, max_row=row_fim)
        c2.add_data(d_sp, titles_from_data=True)
        c2.set_categories(cats_b)
        ws.add_chart(c2, f"A{row_fim + 22}")

    wb.save(caminho)
    print(f"\n[OK] Planilha salva em:\n   {caminho}")
    return caminho


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    ANOS = range(ANO_INICIAL, ANO_FINAL + 1)

    print("=" * 60)
    print("[CVM] Baixando dados da CVM (DRE + BPP -- 1 request por ano)...")
    dre_total, bpp_total, anos_baixados = baixar_dados_cvm(ANOS)

    if not anos_baixados:
        raise RuntimeError("Nenhum ano da CVM foi baixado com sucesso.")

    print(f"[OK] Dados disponiveis: {anos_baixados[0]}-{anos_baixados[-1]}")

    print("\n[BCB] Baixando SELIC...")
    df_selic = baixar_selic(ANO_INICIAL, ANO_FINAL)

    print("[YF] Baixando IBOV anual...")
    df_ibov = baixar_ibov_anual(ANO_INICIAL, ANO_FINAL)

    resultados = {}

    for banco in BANCOS:
        print(f"\n{'='*60}")
        print(f"[BANCO] Processando: {banco['nome']}")

        df_lucro = extrair_lucro(dre_total, banco["cnpj"])
        df_pl    = extrair_pl(bpp_total,    banco["cnpj"])
        df_roe   = calcular_roe(df_lucro, df_pl)

        if df_roe.empty:
            print(f"  [AVISO] Dados insuficientes para {banco['nome']}, pulando.")
            continue

        print(f"  [BETA] Calculando beta -- {banco['ticker']}...")
        beta = calcular_beta(banco["ticker"], ANO_INICIAL, ANO_FINAL)
        print(f"  Beta = {beta:.4f}")

        df_capm = calcular_capm(df_roe, df_selic, df_ibov, beta)
        resultados[banco["nome"]] = df_capm

        print(
            df_capm[["Ano", "ROE", "Ke", "spread_valor"]]
            .assign(
                ROE=lambda x: x["ROE"].map("{:.1%}".format),
                Ke=lambda x: x["Ke"].map("{:.1%}".format),
                spread_valor=lambda x: x["spread_valor"].map("{:.1%}".format),
            )
            .to_string(index=False)
        )

    if resultados:
        print(f"\n{'='*60}")
        print("[EXCEL] Gerando planilha com graficos...")
        criar_excel(resultados, PASTA_SAIDA)
    else:
        print("\n[ERRO] Nenhum banco processado com sucesso.")
